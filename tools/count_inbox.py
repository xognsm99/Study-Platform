from pathlib import Path
import openpyxl as ox

p = Path("data/inbox")
files = sorted(p.glob("*.xlsx"))

print("PWD =", Path(".").resolve())
print("INBOX =", p.resolve())

if not files:
    print("NO XLSX FOUND")
    raise SystemExit(0)

total = 0

for f in files:
    wb = ox.load_workbook(f, data_only=True)
    ws = wb.active

    # 헤더 1행 가정 + 빈 행 제외
    cnt = 0
    max_col = min(ws.max_column, 20)

    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            cnt += 1

    total += cnt
    print(f.name, "=>", cnt)

print("TOTAL_EXPECTED =", total)
